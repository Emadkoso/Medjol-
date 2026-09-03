# app.py
# ============================================================
# MEDJOL FARM MANAGER V3
# بوت ذكي لإدارة العمال واليوميات والفطور والمصاريف والتقارير
# ============================================================

import os
import re
import json
import html
import sqlite3
import logging
import asyncio

from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo

import httpx
import uvicorn
from fastapi import FastAPI
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from weasyprint import HTML


# ============================================================
# CONFIG
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s"
)

logger = logging.getLogger("medjol")

app = FastAPI(title="Medjol Farm Manager V3")

TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

OPENROUTER_MODEL = os.getenv(
    "OPENROUTER_MODEL",
    "meta-llama/llama-3.3-70b-instruct"
)

PORT = int(os.getenv("PORT", "8000"))
DB_NAME = os.getenv("DB_NAME", "farm.db")

# الأردن
TIMEZONE = os.getenv("TIMEZONE", "Asia/Amman")

# ============================================================
# الأجر الثابت
# ============================================================

FIXED_HOURLY_RATE = 1.50

MAX_WORKERS = 100
MAX_HOURS_PER_DAY = 24
MAX_EXPENSE = 1_000_000
MAX_BREAKFAST_PER_WORKER = 100


# ============================================================
# TIME
# ============================================================

def now_local():
    return datetime.now(ZoneInfo(TIMEZONE))


def today_str():
    return now_local().strftime("%Y-%m-%d")


def parse_date_text(text):
    """
    يفهم:
    اليوم
    أمس
    امس
    امبارح
    البارحة
    غداً
    غدا
    بكرة
    وكذلك:
    1/9/2026
    01/09/2026
    2026-09-01
    1-9-2026
    """

    if not text:
        return None

    original = text.strip()
    t = normalize_text(original).lower()

    today = now_local().date()

    # --------------------------------------------------------
    # كلمات التاريخ
    # --------------------------------------------------------

    if any(x in t for x in [
        "اليوم",
        "هذا اليوم"
    ]):
        return today.isoformat()

    if any(x in t for x in [
        "امس",
        "امبارح",
        "البارحة",
        "مبارح"
    ]):
        return (today - timedelta(days=1)).isoformat()

    if any(x in t for x in [
        "غدا",
        "بكره",
        "بكرة",
        "غدًا"
    ]):
        return (today + timedelta(days=1)).isoformat()

    # --------------------------------------------------------
    # YYYY-MM-DD
    # --------------------------------------------------------

    match = re.search(
        r"(?<!\d)(20\d{2})[-/](\d{1,2})[-/](\d{1,2})(?!\d)",
        t
    )

    if match:
        try:
            d = date(
                int(match.group(1)),
                int(match.group(2)),
                int(match.group(3))
            )
            return d.isoformat()
        except ValueError:
            return None

    # --------------------------------------------------------
    # DD/MM/YYYY أو DD-MM-YYYY
    # --------------------------------------------------------

    match = re.search(
        r"(?<!\d)(\d{1,2})[-/](\d{1,2})[-/](20\d{2})(?!\d)",
        t
    )

    if match:
        try:
            d = date(
                int(match.group(3)),
                int(match.group(2)),
                int(match.group(1))
            )
            return d.isoformat()
        except ValueError:
            return None

    # --------------------------------------------------------
    # DD/MM بدون سنة
    # --------------------------------------------------------

    match = re.search(
        r"(?<!\d)(\d{1,2})[-/](\d{1,2})(?!\d)",
        t
    )

    if match:
        try:
            day = int(match.group(1))
            month = int(match.group(2))

            d = date(
                today.year,
                month,
                day
            )

            return d.isoformat()

        except ValueError:
            return None

    return None


# ============================================================
# DATABASE
# ============================================================

def db():
    conn = sqlite3.connect(
        DB_NAME,
        timeout=30
    )

    conn.row_factory = sqlite3.Row

    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    return conn


def init_db():

    conn = db()
    c = conn.cursor()

    # --------------------------------------------------------
    # العمال
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS workers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE COLLATE NOCASE,
            phone TEXT,
            hourly_rate REAL NOT NULL DEFAULT 1.5,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL
        )
    """)

    # --------------------------------------------------------
    # اليوميات
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS daily_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            notes TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    # --------------------------------------------------------
    # ساعات العمال
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS work_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            daily_log_id INTEGER NOT NULL,
            worker_id INTEGER NOT NULL,
            hours REAL NOT NULL,
            task TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            UNIQUE(daily_log_id, worker_id),
            FOREIGN KEY(daily_log_id)
                REFERENCES daily_logs(id)
                ON DELETE CASCADE,
            FOREIGN KEY(worker_id)
                REFERENCES workers(id)
        )
    """)

    # --------------------------------------------------------
    # المصاريف
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            daily_log_id INTEGER NOT NULL,
            category TEXT DEFAULT 'عام',
            amount REAL NOT NULL,
            notes TEXT DEFAULT '',
            FOREIGN KEY(daily_log_id)
                REFERENCES daily_logs(id)
                ON DELETE CASCADE
        )
    """)

    # --------------------------------------------------------
    # جلسات الحوار
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            chat_id INTEGER PRIMARY KEY,
            step TEXT NOT NULL,
            data TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    # --------------------------------------------------------
    # الإعدادات
    # --------------------------------------------------------

    c.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)

    # --------------------------------------------------------
    # تثبيت أجر الساعة على 1.5
    # --------------------------------------------------------

    c.execute("""
        INSERT INTO settings(key, value)
        VALUES ('default_hourly_rate', '1.5')
        ON CONFLICT(key)
        DO UPDATE SET value = '1.5'
    """)

    # تحديث أجر العمال المسجلين إلى السعر الثابت
    c.execute("""
        UPDATE workers
        SET hourly_rate = 1.5
    """)

    conn.commit()
    conn.close()


init_db()


# ============================================================
# SETTINGS
# ============================================================

def default_hourly_rate():
    # ثابت وغير قابل للتغيير من المستخدم
    return FIXED_HOURLY_RATE


# ============================================================
# TELEGRAM
# ============================================================

async def telegram_request(method, **kwargs):

    if not TELEGRAM_BOT_TOKEN:
        logger.error("TELEGRAM_BOT_TOKEN مفقود")
        return None

    url = (
        f"https://api.telegram.org/"
        f"bot{TELEGRAM_BOT_TOKEN}/{method}"
    )

    try:

        async with httpx.AsyncClient() as client:

            response = await client.post(
                url,
                **kwargs,
                timeout=30
            )

            if response.status_code != 200:

                logger.error(
                    "Telegram %s error: %s",
                    method,
                    response.text
                )

                return None

            body = response.json()

            if not body.get("ok"):
                logger.error(
                    "Telegram API error: %s",
                    body
                )
                return None

            return body

    except Exception as e:

        logger.exception(
            "Telegram request failed: %s",
            e
        )

        return None


async def send_message(chat_id, text):

    if not text:
        return None

    # Telegram maximum is 4096.
    # نستخدم 3900 كهامش أمان.
    chunks = []

    while len(text) > 3900:

        split_at = text.rfind(
            "\n",
            0,
            3900
        )

        if split_at < 1000:
            split_at = 3900

        chunks.append(
            text[:split_at]
        )

        text = text[split_at:].lstrip()

    if text:
        chunks.append(text)

    result = None

    for chunk in chunks:

        result = await telegram_request(
            "sendMessage",
            json={
                "chat_id": chat_id,
                "text": chunk
            }
        )

    return result


async def send_document(
    chat_id,
    filename,
    content,
    caption=""
):

    if not TELEGRAM_BOT_TOKEN:
        return None

    url = (
        f"https://api.telegram.org/"
        f"bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    )

    try:

        async with httpx.AsyncClient() as client:

            response = await client.post(
                url,
                data={
                    "chat_id": str(chat_id),
                    "caption": caption
                },
                files={
                    "document": (
                        filename,
                        content,
                        "application/octet-stream"
                    )
                },
                timeout=60
            )

            if response.status_code != 200:

                logger.error(
                    "Document error: %s",
                    response.text
                )

                return None

            body = response.json()

            if not body.get("ok"):
                logger.error(
                    "Telegram document error: %s",
                    body
                )
                return None

            return body

    except Exception as e:

        logger.exception(
            "Document error: %s",
            e
        )

        return None


# ============================================================
# TEXT NORMALIZATION
# ============================================================

def normalize_text(text):

    if not text:
        return ""

    text = str(text).strip()

    translation = str.maketrans(
        "٠١٢٣٤٥٦٧٨٩",
        "0123456789"
    )

    text = text.translate(
        translation
    )

    text = text.replace(
        "٫",
        "."
    )

    text = text.replace(
        "،",
        ","
    )

    text = text.replace(
        "؛",
        ";"
    )

    # توحيد بعض الحروف للمقارنة فقط
    text = text.replace(
        "أ",
        "ا"
    )

    text = text.replace(
        "إ",
        "ا"
    )

    text = text.replace(
        "آ",
        "ا"
    )

    return text


def normalize_name(name):

    if not name:
        return ""

    name = str(name).strip()

    # لا نريد تغيير أسماء الأشخاص الأصلية
    name = re.sub(
        r"\s+",
        " ",
        name
    )

    name = name.strip(
        " ,.;:،؛-–—"
    )

    return name


# ============================================================
# NUMBERS
# ============================================================

ARABIC_NUMBERS = {
    "صفر": 0,
    "واحد": 1,
    "واحدة": 1,
    "اثنان": 2,
    "اثنين": 2,
    "اثنتان": 2,
    "ثلاثة": 3,
    "ثلاث": 3,
    "اربعة": 4,
    "أربعة": 4,
    "اربع": 4,
    "أربع": 4,
    "خمسة": 5,
    "خمس": 5,
    "ستة": 6,
    "ست": 6,
    "سبعة": 7,
    "سبع": 7,
    "ثمانية": 8,
    "ثماني": 8,
    "ثمان": 8,
    "تسعة": 9,
    "تسع": 9,
    "عشرة": 10,
    "عشر": 10,
    "عشرون": 20,
    "ثلاثون": 30,
    "اربعون": 40,
    "أربعون": 40,
    "خمسون": 50,
    "ستون": 60,
    "سبعون": 70,
    "ثمانون": 80,
    "تسعون": 90,
    "مئة": 100,
    "مائة": 100
}


def number_from_text(text):

    if not text:
        return None

    original = str(text)

    normalized = normalize_text(
        original
    ).lower()

    # --------------------------------------------------------
    # رقم عشري
    # --------------------------------------------------------

    match = re.search(
        r"(?<!\d)(\d+(?:[.,]\d+)?)(?!\d)",
        normalized
    )

    if match:

        try:

            return float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

        except Exception:
            pass

    # --------------------------------------------------------
    # رقم عربي بالكلمات
    # --------------------------------------------------------

    words = normalized.split()

    for word in words:

        word = word.strip(
            ".,:;()[]{}"
        )

        if word in ARABIC_NUMBERS:
            return float(
                ARABIC_NUMBERS[word]
            )

    return None


# ============================================================
# YES / NO
# ============================================================

def is_no(text):

    t = normalize_text(
        text
    ).lower().strip()

    return t in {
        "لا",
        "لا يوجد",
        "لا يوجد شيء",
        "ما في",
        "مفيش",
        "بدون",
        "0",
        "صفر",
        "لا يوجد مصاريف",
        "بدون مصاريف",
        "بدون فطور",
        "لا يوجد فطور",
        "ما في فطور"
    }


def is_yes(text):

    t = normalize_text(
        text
    ).lower().strip()

    return t in {
        "نعم",
        "ايوه",
        "ايوا",
        "اه",
        "آه",
        "موافق",
        "تأكيد",
        "تأكيد الحفظ",
        "اكد",
        "احفظ",
        "حفظ",
        "صحيح",
        "نعم احفظ"
    }


# ============================================================
# SESSION
# ============================================================

def get_session(chat_id):

    conn = db()

    row = conn.execute(
        """
        SELECT step, data, updated_at
        FROM sessions
        WHERE chat_id = ?
        """,
        (chat_id,)
    ).fetchone()

    conn.close()

    if not row:
        return None

    try:

        # الجلسة تنتهي بعد 6 ساعات
        updated = datetime.fromisoformat(
            row["updated_at"]
        )

        age = (
            now_local()
            - updated
        ).total_seconds()

        if age > 21600:

            clear_session(chat_id)
            return None

        return {
            "step": row["step"],
            "data": json.loads(
                row["data"]
            )
        }

    except Exception:

        clear_session(chat_id)
        return None


def save_session(
    chat_id,
    step,
    data
):

    conn = db()

    conn.execute(
        """
        INSERT INTO sessions(
            chat_id,
            step,
            data,
            updated_at
        )
        VALUES (?, ?, ?, ?)

        ON CONFLICT(chat_id)
        DO UPDATE SET
            step = excluded.step,
            data = excluded.data,
            updated_at = excluded.updated_at
        """,
        (
            chat_id,
            step,
            json.dumps(
                data,
                ensure_ascii=False
            ),
            now_local().isoformat()
        )
    )

    conn.commit()
    conn.close()


def clear_session(chat_id):

    conn = db()

    conn.execute(
        "DELETE FROM sessions WHERE chat_id = ?",
        (chat_id,)
    )

    conn.commit()
    conn.close()


# ============================================================
# VALIDATION
# ============================================================

def validate_worker_name(name):

    name = normalize_name(name)

    if not name:
        return False

    if len(name) < 2:
        return False

    if len(name) > 100:
        return False

    if len(name.split()) > 6:
        return False

    return True


def clean_names(names):

    if not isinstance(names, list):
        return []

    result = []

    for name in names:

        if not isinstance(name, str):
            continue

        name = normalize_name(name)

        if not validate_worker_name(name):
            continue

        if name.lower() not in [
            x.lower()
            for x in result
        ]:
            result.append(name)

    return result


def validate_amount(
    value,
    maximum=MAX_EXPENSE
):

    try:

        value = float(value)

    except Exception:

        return False

    return (
        0 <= value <= maximum
    )


# ============================================================
# AI EXTRACTION
# ============================================================

AI_SYSTEM_PROMPT = """
أنت محرك استخراج بيانات فقط لنظام إدارة مزرعة.

مهمتك استخراج المعلومات الموجودة صراحة في رسالة المستخدم.

ممنوع التخمين.
ممنوع اختراع أسماء.
ممنوع اختراع أرقام.
ممنوع تغيير التاريخ.
ممنوع اعتبار كلمات مثل عامل، ساعة، دينار، فطور، مصاريف أسماء أشخاص.

أجرة الساعة ليست من البيانات التي يستخرجها النظام:
أجرة الساعة ثابتة دائماً = 1.50 دينار أردني.
لا تحاول تغييرها حتى لو طلب المستخدم ذلك.

DATE:
إذا ذكر المستخدم:
- اليوم => تاريخ اليوم
- أمس / امبارح / مبارح / البارحة => أمس
- غداً / بكرة => غداً
- تاريخ صريح => حوّله إلى YYYY-MM-DD
إذا لم يذكر تاريخاً => date = null.

WORKERS:
إذا قال "3 عمال" => workers_count = 3.
إذا قال "أحمد ومحمد وخالد" => workers_names = ["أحمد","محمد","خالد"] و workers_count = 3.
إذا لم يذكر الأسماء => workers_names = [].

HOURS:
total_hours = ساعات العمل لكل عامل إذا كانت الساعات متساوية.
إذا كانت مختلفة بين العمال، استخرجها فقط إذا أمكن تحديدها بوضوح.

BREAKFAST:
هناك فرق مهم:

"فطور لكل عامل دينار"
=> breakfast_per_worker = 1

"كل عامل له دينار فطور"
=> breakfast_per_worker = 1

"فطور لكل عامل 1.5"
=> breakfast_per_worker = 1.5

أما:
"فطور 6 دنانير"
=> breakfast_total = 6

إذا لم يكن واضحاً هل المبلغ لكل عامل أو إجمالي:
لا تخمن، استخدم null.

OTHER EXPENSES:
"مصاريف 10 دنانير"
=> other_expenses = 10

لا تعتبر الفطور ضمن other_expenses إذا تم ذكره بشكل مستقل.

NOTES:
ضع المعلومات الإضافية المهمة فقط.

أعد JSON فقط، بدون أي شرح.

الصيغة:

{
  "date": "YYYY-MM-DD"|null,
  "workers_count": number|null,
  "workers_names": ["name"],
  "total_hours": number|null,
  "breakfast_per_worker": number|null,
  "breakfast_total": number|null,
  "other_expenses": number|null,
  "notes": string|null
}
"""


def extract_json_content(content):

    if not content:
        return None

    content = content.strip()

    # إزالة markdown fences
    content = re.sub(
        r"^```(?:json)?",
        "",
        content,
        flags=re.IGNORECASE
    )

    content = re.sub(
        r"```$",
        "",
        content
    ).strip()

    # استخراج أول JSON
    start = content.find("{")
    end = content.rfind("}")

    if start == -1 or end == -1:
        return None

    content = content[
        start:end + 1
    ]

    try:
        return json.loads(content)
    except Exception:
        return None


def normalize_ai_data(data):

    if not isinstance(data, dict):
        return None

    result = {
        "date": data.get("date"),
        "workers_count": data.get(
            "workers_count"
        ),
        "workers_names": clean_names(
            data.get(
                "workers_names",
                []
            )
        ),
        "total_hours": data.get(
            "total_hours"
        ),
        "breakfast_per_worker": data.get(
            "breakfast_per_worker"
        ),
        "breakfast_total": data.get(
            "breakfast_total"
        ),
        "other_expenses": data.get(
            "other_expenses"
        ),
        "notes": data.get(
            "notes"
        ) or ""
    }

    # --------------------------------------------------------
    # التاريخ
    # --------------------------------------------------------

    if result["date"]:

        parsed = parse_date_text(
            str(result["date"])
        )

        if parsed:
            result["date"] = parsed
        else:
            # ربما أعاد AI تاريخاً صحيحاً بصيغة YYYY-MM-DD
            try:
                result["date"] = date.fromisoformat(
                    str(result["date"])
                ).isoformat()
            except Exception:
                result["date"] = None

    # --------------------------------------------------------
    # العدد
    # --------------------------------------------------------

    if result["workers_count"] is not None:

        try:

            result["workers_count"] = int(
                float(
                    result["workers_count"]
                )
            )

        except Exception:

            result["workers_count"] = None

    # --------------------------------------------------------
    # الساعات
    # --------------------------------------------------------

    if result["total_hours"] is not None:

        try:

            result["total_hours"] = float(
                result["total_hours"]
            )

        except Exception:

            result["total_hours"] = None

    # --------------------------------------------------------
    # الفطور لكل عامل
    # --------------------------------------------------------

    if result["breakfast_per_worker"] is not None:

        try:

            result[
                "breakfast_per_worker"
            ] = float(
                result[
                    "breakfast_per_worker"
                ]
            )

        except Exception:

            result[
                "breakfast_per_worker"
            ] = None

    # --------------------------------------------------------
    # إجمالي الفطور
    # --------------------------------------------------------

    if result["breakfast_total"] is not None:

        try:

            result["breakfast_total"] = float(
                result["breakfast_total"]
            )

        except Exception:

            result["breakfast_total"] = None

    # --------------------------------------------------------
    # المصاريف العامة
    # --------------------------------------------------------

    if result["other_expenses"] is not None:

        try:

            result["other_expenses"] = float(
                result["other_expenses"]
            )

        except Exception:

            result["other_expenses"] = None

    # --------------------------------------------------------
    # validation
    # --------------------------------------------------------

    if result["workers_count"] is not None:

        if not (
            1 <= result["workers_count"]
            <= MAX_WORKERS
        ):
            return {
                "_validation_error":
                    f"عدد العمال يجب أن يكون بين 1 و{MAX_WORKERS}."
            }

    if result["total_hours"] is not None:

        if not (
            0 <= result["total_hours"]
            <= MAX_HOURS_PER_DAY
        ):
            return {
                "_validation_error":
                    "ساعات العمل يجب أن تكون بين 0 و24."
            }

    if result["breakfast_per_worker"] is not None:

        if not validate_amount(
            result["breakfast_per_worker"],
            MAX_BREAKFAST_PER_WORKER
        ):
            return {
                "_validation_error":
                    "قيمة فطور العامل غير صحيحة."
            }

    if result["breakfast_total"] is not None:

        if not validate_amount(
            result["breakfast_total"]
        ):
            return {
                "_validation_error":
                    "إجمالي الفطور غير صحيح."
            }

    if result["other_expenses"] is not None:

        if not validate_amount(
            result["other_expenses"]
        ):
            return {
                "_validation_error":
                    "قيمة المصاريف غير صحيحة."
            }

    # --------------------------------------------------------
    # consistency
    # --------------------------------------------------------

    if (
        result["workers_count"]
        and result["workers_names"]
    ):

        if (
            len(result["workers_names"])
            != result["workers_count"]
        ):

            return {
                "_validation_error":
                    f"ذكرت {result['workers_count']} عمال "
                    f"لكن وجدت {len(result['workers_names'])} أسماء. "
                    "لن أخمّن العدد الصحيح."
            }

    return result


async def extract_data_with_ai(text):

    if not OPENROUTER_API_KEY:
        return None

    text = normalize_text(text)

    headers = {
        "Authorization":
            f"Bearer {OPENROUTER_API_KEY}",

        "Content-Type":
            "application/json",

        "HTTP-Referer":
            "https://medjol.onrender.com",

        "X-Title":
            "Medjol Farm Manager"
    }

    payload = {
        "model": OPENROUTER_MODEL,

        "messages": [
            {
                "role": "system",
                "content":
                    AI_SYSTEM_PROMPT
            },
            {
                "role": "user",
                "content":
                    text
            }
        ],

        "temperature": 0,

        "response_format": {
            "type": "json_object"
        }
    }

    try:

        async with httpx.AsyncClient() as client:

            response = await client.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30
            )

            # بعض النماذج قد لا تدعم response_format
            if response.status_code == 400:

                payload.pop(
                    "response_format",
                    None
                )

                response = await client.post(
                    "https://openrouter.ai/api/v1/chat/completions",
                    headers=headers,
                    json=payload,
                    timeout=30
                )

        if response.status_code != 200:

            logger.error(
                "OpenRouter error %s",
                response.status_code
            )

            return None

        body = response.json()

        content = (
            body
            .get("choices", [{}])[0]
            .get("message", {})
            .get("content")
        )

        data = extract_json_content(
            content
        )

        if not data:
            return None

        return normalize_ai_data(
            data
        )

    except Exception as e:

        logger.exception(
            "AI extraction failed: %s",
            e
        )

        return None


# ============================================================
# DETERMINISTIC FALLBACK
# ============================================================

def deterministic_extract(text):

    t = normalize_text(
        text
    )

    result = {
        "date": parse_date_text(text),
        "workers_count": None,
        "workers_names": [],
        "total_hours": None,
        "breakfast_per_worker": None,
        "breakfast_total": None,
        "other_expenses": None,
        "notes": ""
    }

    # --------------------------------------------------------
    # عدد العمال
    # --------------------------------------------------------

    match = re.search(
        r"(\d+(?:\.\d+)?)\s*(?:عامل|عمال|عاملين|شخص|اشخاص)",
        t,
        flags=re.IGNORECASE
    )

    if match:

        result["workers_count"] = int(
            float(match.group(1))
        )

    # --------------------------------------------------------
    # الساعات
    # --------------------------------------------------------

    match = re.search(
        r"(\d+(?:[.,]\d+)?)\s*(?:ساعة|ساعات)",
        t,
        flags=re.IGNORECASE
    )

    if match:

        result["total_hours"] = float(
            match.group(1).replace(
                ",",
                "."
            )
        )

    # --------------------------------------------------------
    # فطور لكل عامل
    # --------------------------------------------------------

    breakfast_per_worker_patterns = [
        r"فطور\s+لكل\s+عامل\s+(\d+(?:[.,]\d+)?)",
        r"فطور\s+للعامل\s+(\d+(?:[.,]\d+)?)",
        r"الفطور\s+لكل\s+عامل\s+(\d+(?:[.,]\d+)?)",
        r"كل\s+عامل\s+.*?فطور.*?(\d+(?:[.,]\d+)?)",
    ]

    for pattern in breakfast_per_worker_patterns:

        match = re.search(
            pattern,
            t,
            flags=re.IGNORECASE
        )

        if match:

            result[
                "breakfast_per_worker"
            ] = float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

            break

    # --------------------------------------------------------
    # فطور إجمالي
    # --------------------------------------------------------

    if result[
        "breakfast_per_worker"
    ] is None:

        match = re.search(
            r"(?:فطور|الفطور)\s*(?:=|:)?\s*(\d+(?:[.,]\d+)?)\s*(?:دينار|دنانير|د\.أ)?",
            t,
            flags=re.IGNORECASE
        )

        if match:

            result[
                "breakfast_total"
            ] = float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

    # --------------------------------------------------------
    # مصاريف عامة
    # --------------------------------------------------------

    expense_patterns = [
        r"(?:مصاريف|المصاريف|مصروف|صرف|دفعت)\s*(?:=|:)?\s*(\d+(?:[.,]\d+)?)",
        r"مصاريف\s+(\d+(?:[.,]\d+)?)"
    ]

    for pattern in expense_patterns:

        match = re.search(
            pattern,
            t,
            flags=re.IGNORECASE
        )

        if match:

            result[
                "other_expenses"
            ] = float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

            break

    # --------------------------------------------------------
    # أسماء العمال
    # --------------------------------------------------------

    names_match = re.search(
        r"(?:عمال|العاملين|اسماء العمال|أسماء العمال)\s*[:=]\s*(.+)",
        text,
        flags=re.IGNORECASE
    )

    if names_match:

        names_text = names_match.group(1)

        # إيقاف النص عند الساعات أو المصاريف
        names_text = re.split(
            r"\s+(?:كل|ساعة|ساعات|المصاريف|مصاريف|فطور|الفطور)\b",
            names_text,
            maxsplit=1,
            flags=re.IGNORECASE
        )[0]

        parts = re.split(
            r"[,،;؛\n]+|\s+و\s+",
            names_text
        )

        result[
            "workers_names"
        ] = clean_names(parts)

    # --------------------------------------------------------
    # إذا لم توجد أسماء بعد
    # نحاول جملة "اشتغل أحمد ومحمد"
    # --------------------------------------------------------

    if not result["workers_names"]:

        match = re.search(
            r"(?:اشتغل|اشتغلوا|حضر|حضروا)\s+(.+?)(?:\s+\d+\s*(?:ساعة|ساعات)|\s+مصاريف|\s+فطور|$)",
            text,
            flags=re.IGNORECASE
        )

        if match:

            names_text = match.group(1)

            parts = re.split(
                r"[,،;؛]+|\s+و\s+",
                names_text
            )

            result[
                "workers_names"
            ] = clean_names(parts)

    if (
        result["workers_count"] is None
        and result["workers_names"]
    ):
        result["workers_count"] = len(
            result["workers_names"]
        )

    return normalize_ai_data(
        result
    )


# ============================================================
# WORKERS
# ============================================================

def get_worker_by_name(name):

    name = normalize_name(name)

    conn = db()

    row = conn.execute(
        """
        SELECT *
        FROM workers
        WHERE name = ? COLLATE NOCASE
        """,
        (name,)
    ).fetchone()

    conn.close()

    return row


def create_worker(name):

    name = normalize_name(name)

    if not validate_worker_name(name):
        return None

    conn = db()

    try:

        cursor = conn.execute(
            """
            INSERT INTO workers(
                name,
                hourly_rate,
                active,
                created_at
            )
            VALUES (?, 1.5, 1, ?)
            """,
            (
                name,
                now_local().isoformat()
            )
        )

        conn.commit()

        return cursor.lastrowid

    except sqlite3.IntegrityError:

        row = conn.execute(
            """
            SELECT id
            FROM workers
            WHERE name = ? COLLATE NOCASE
            """,
            (name,)
        ).fetchone()

        return (
            row["id"]
            if row
            else None
        )

    finally:

        conn.close()


def get_active_workers():

    conn = db()

    rows = conn.execute(
        """
        SELECT *
        FROM workers
        WHERE active = 1
        ORDER BY name COLLATE NOCASE
        """
    ).fetchall()

    conn.close()

    return rows


def deactivate_worker(name):

    conn = db()

    cursor = conn.execute(
        """
        UPDATE workers
        SET active = 0
        WHERE name = ? COLLATE NOCASE
        """,
        (normalize_name(name),)
    )

    conn.commit()

    changed = cursor.rowcount > 0

    conn.close()

    return changed


# ============================================================
# DAILY LOGS
# ============================================================

def get_daily_log(date_value):

    conn = db()

    daily = conn.execute(
        """
        SELECT *
        FROM daily_logs
        WHERE date = ?
        """,
        (date_value,)
    ).fetchone()

    if not daily:

        conn.close()
        return None

    workers = conn.execute(
        """
        SELECT
            w.id,
            w.name,
            1.5 AS hourly_rate,
            wl.hours,
            wl.task,
            wl.notes
        FROM work_logs wl
        JOIN workers w
            ON w.id = wl.worker_id
        WHERE wl.daily_log_id = ?
        ORDER BY w.name COLLATE NOCASE
        """,
        (daily["id"],)
    ).fetchall()

    expenses = conn.execute(
        """
        SELECT *
        FROM expenses
        WHERE daily_log_id = ?
        ORDER BY id
        """,
        (daily["id"],)
    ).fetchall()

    conn.close()

    return {
        "id": daily["id"],
        "date": daily["date"],
        "notes": daily["notes"] or "",
        "workers": workers,
        "expenses": expenses
    }


def daily_exists(date_value):

    conn = db()

    row = conn.execute(
        """
        SELECT id
        FROM daily_logs
        WHERE date = ?
        """,
        (date_value,)
    ).fetchone()

    conn.close()

    return row is not None


def save_daily_data(data):

    """
    الحفظ داخل transaction واحدة.
    أجرة الساعة دائماً 1.50 د.أ.
    """

    date_value = data["date"]

    workers = data["workers"]

    expenses = data.get(
        "expenses",
        []
    )

    notes = data.get(
        "notes",
        ""
    )

    conn = db()

    try:

        conn.execute("BEGIN")

        existing = conn.execute(
            """
            SELECT id
            FROM daily_logs
            WHERE date = ?
            """,
            (date_value,)
        ).fetchone()

        if existing:

            raise ValueError(
                f"هناك يومية موجودة بالفعل بتاريخ {date_value}"
            )

        cursor = conn.execute(
            """
            INSERT INTO daily_logs(
                date,
                notes,
                created_at,
                updated_at
            )
            VALUES (?, ?, ?, ?)
            """,
            (
                date_value,
                notes,
                now_local().isoformat(),
                now_local().isoformat()
            )
        )

        daily_id = cursor.lastrowid

        # ----------------------------------------------------
        # العمال
        # ----------------------------------------------------

        for worker in workers:

            name = normalize_name(
                worker["name"]
            )

            hours = float(
                worker["hours"]
            )

            worker_row = conn.execute(
                """
                SELECT id
                FROM workers
                WHERE name = ? COLLATE NOCASE
                """,
                (name,)
            ).fetchone()

            if worker_row:

                worker_id = worker_row["id"]

            else:

                cursor = conn.execute(
                    """
                    INSERT INTO workers(
                        name,
                        hourly_rate,
                        active,
                        created_at
                    )
                    VALUES (?, 1.5, 1, ?)
                    """,
                    (
                        name,
                        now_local().isoformat()
                    )
                )

                worker_id = cursor.lastrowid

            conn.execute(
                """
                INSERT INTO work_logs(
                    daily_log_id,
                    worker_id,
                    hours,
                    task,
                    notes
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    daily_id,
                    worker_id,
                    hours,
                    worker.get(
                        "task",
                        ""
                    ),
                    worker.get(
                        "notes",
                        ""
                    )
                )
            )

        # ----------------------------------------------------
        # المصاريف
        # ----------------------------------------------------

        for expense in expenses:

            amount = float(
                expense["amount"]
            )

            if amount < 0:
                raise ValueError(
                    "لا يمكن أن تكون المصاريف سالبة."
                )

            conn.execute(
                """
                INSERT INTO expenses(
                    daily_log_id,
                    category,
                    amount,
                    notes
                )
                VALUES (?, ?, ?, ?)
                """,
                (
                    daily_id,
                    expense.get(
                        "category",
                        "عام"
                    ),
                    amount,
                    expense.get(
                        "notes",
                        ""
                    )
                )
            )

        conn.commit()

        return daily_id

    except Exception:

        conn.rollback()
        raise

    finally:

        conn.close()


# ============================================================
# CALCULATIONS
# ============================================================

def calculate_daily_total(log):

    wages = 0.0

    breakfast = 0.0
    other_expenses = 0.0

    for worker in log["workers"]:

        hours = float(
            worker["hours"]
        )

        # السعر ثابت
        wages += (
            hours
            * FIXED_HOURLY_RATE
        )

    for expense in log["expenses"]:

        amount = float(
            expense["amount"]
        )

        category = (
            expense["category"]
            or ""
        ).strip().lower()

        if category in [
            "فطور",
            "الفطور"
        ]:
            breakfast += amount

        else:
            other_expenses += amount

    total_expenses = (
        breakfast
        + other_expenses
    )

    total = (
        wages
        + total_expenses
    )

    return (
        wages,
        breakfast,
        other_expenses,
        total
    )


# ============================================================
# DISPLAY
# ============================================================

def format_money(value):

    return (
        f"{float(value):.2f} د.أ"
    )


def format_daily_summary(log):

    (
        wages,
        breakfast,
        other_expenses,
        total
    ) = calculate_daily_total(log)

    lines = []

    lines.append(
        f"📋 اليومية: {log['date']}"
    )

    lines.append("")

    lines.append(
        f"👷 عدد العمال: "
        f"{len(log['workers'])}"
    )

    if log["workers"]:

        lines.append("")

        for index, worker in enumerate(
            log["workers"],
            1
        ):

            hours = float(
                worker["hours"]
            )

            wage = (
                hours
                * FIXED_HOURLY_RATE
            )

            lines.append(
                f"{index}. {worker['name']} — "
                f"{hours:g} ساعة × "
                f"{FIXED_HOURLY_RATE:.2f} = "
                f"{format_money(wage)}"
            )

    lines.append("")

    lines.append(
        f"⏱️ أجر الساعة الثابت: "
        f"{FIXED_HOURLY_RATE:.2f} د.أ"
    )

    lines.append(
        f"💵 أجور العمال: "
        f"{format_money(wages)}"
    )

    lines.append(
        f"🍳 الفطور: "
        f"{format_money(breakfast)}"
    )

    lines.append(
        f"💰 المصاريف الأخرى: "
        f"{format_money(other_expenses)}"
    )

    lines.append(
        f"🏆 الإجمالي: "
        f"{format_money(total)}"
    )

    if log["expenses"]:

        lines.append("")
        lines.append("🧾 تفاصيل المصاريف:")

        for expense in log["expenses"]:

            lines.append(
                f"- {expense['category']}: "
                f"{format_money(expense['amount'])}"
                + (
                    f" ({expense['notes']})"
                    if expense["notes"]
                    else ""
                )
            )

    if log["notes"]:

        lines.append("")
        lines.append(
            f"📝 ملاحظات: "
            f"{log['notes']}"
        )

    return "\n".join(lines)


# ============================================================
# REPORT DATA
# ============================================================

def get_logs_between(
    date_from,
    date_to
):

    conn = db()

    daily_rows = conn.execute(
        """
        SELECT *
        FROM daily_logs
        WHERE date BETWEEN ? AND ?
        ORDER BY date
        """,
        (
            date_from,
            date_to
        )
    ).fetchall()

    result = []

    for daily in daily_rows:

        workers = conn.execute(
            """
            SELECT
                w.name,
                1.5 AS hourly_rate,
                wl.hours,
                wl.task
            FROM work_logs wl
            JOIN workers w
                ON w.id = wl.worker_id
            WHERE wl.daily_log_id = ?
            ORDER BY w.name COLLATE NOCASE
            """,
            (daily["id"],)
        ).fetchall()

        expenses = conn.execute(
            """
            SELECT
                category,
                amount,
                notes
            FROM expenses
            WHERE daily_log_id = ?
            ORDER BY id
            """
        ).fetchall()

        # تصحيح الاستعلام السابق بإعادة جلب المصاريف لليومية
        expenses = conn.execute(
            """
            SELECT
                category,
                amount,
                notes
            FROM expenses
            WHERE daily_log_id = ?
            ORDER BY id
            """,
            (daily["id"],)
        ).fetchall()

        result.append({
            "date": daily["date"],
            "notes": daily["notes"] or "",
            "workers": workers,
            "expenses": expenses
        })

    conn.close()

    return result


# ============================================================
# TEXT REPORT
# ============================================================

def generate_text_report(
    logs,
    title="📊 التقرير"
):

    if not logs:
        return (
            "لا توجد بيانات في الفترة المطلوبة."
        )

    lines = [
        title,
        "=" * 40
    ]

    grand_wages = 0.0
    grand_breakfast = 0.0
    grand_other_expenses = 0.0

    for log in logs:

        wages = sum(
            float(w["hours"])
            * FIXED_HOURLY_RATE
            for w in log["workers"]
        )

        breakfast = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            in ["فطور", "الفطور"]
        )

        other_expenses = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            not in ["فطور", "الفطور"]
        )

        total = (
            wages
            + breakfast
            + other_expenses
        )

        grand_wages += wages
        grand_breakfast += breakfast
        grand_other_expenses += (
            other_expenses
        )

        lines.append("")

        lines.append(
            f"📅 التاريخ: {log['date']}"
        )

        lines.append(
            f"👷 العمال: "
            f"{len(log['workers'])}"
        )

        lines.append(
            f"⏱️ مجموع الساعات: "
            f"{sum(float(w['hours']) for w in log['workers']):.2f}"
        )

        lines.append(
            f"💵 الأجور: "
            f"{format_money(wages)}"
        )

        lines.append(
            f"🍳 الفطور: "
            f"{format_money(breakfast)}"
        )

        lines.append(
            f"💰 المصاريف الأخرى: "
            f"{format_money(other_expenses)}"
        )

        lines.append(
            f"🏆 إجمالي اليوم: "
            f"{format_money(total)}"
        )

    final_total = (
        grand_wages
        + grand_breakfast
        + grand_other_expenses
    )

    lines.append("")

    lines.append(
        "=" * 40
    )

    lines.append(
        f"💵 إجمالي الأجور: "
        f"{format_money(grand_wages)}"
    )

    lines.append(
        f"🍳 إجمالي الفطور: "
        f"{format_money(grand_breakfast)}"
    )

    lines.append(
        f"💰 إجمالي المصاريف الأخرى: "
        f"{format_money(grand_other_expenses)}"
    )

    lines.append(
        f"🏆 الإجمالي النهائي: "
        f"{format_money(final_total)}"
    )

    return "\n".join(lines)


# ============================================================
# EXCEL
# ============================================================

def generate_excel_report(logs):

    wb = Workbook()

    ws = wb.active
    ws.title = "اليوميات"
    ws.sheet_view.rightToLeft = True

    headers = [
        "التاريخ",
        "العامل",
        "الساعات",
        "أجر الساعة",
        "أجر العامل",
        "ملاحظات"
    ]

    ws.append(headers)

    for cell in ws[1]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for log in logs:

        for worker in log["workers"]:

            hours = float(
                worker["hours"]
            )

            wage = (
                hours
                * FIXED_HOURLY_RATE
            )

            ws.append([
                log["date"],
                worker["name"],
                hours,
                FIXED_HOURLY_RATE,
                wage,
                log["notes"]
            ])

    # --------------------------------------------------------
    # المصاريف
    # --------------------------------------------------------

    expense_ws = wb.create_sheet(
        "المصاريف"
    )

    expense_ws.sheet_view.rightToLeft = True

    expense_ws.append([
        "التاريخ",
        "التصنيف",
        "المبلغ",
        "ملاحظات"
    ])

    for cell in expense_ws[1]:

        cell.font = Font(
            bold=True
        )

    for log in logs:

        for expense in log["expenses"]:

            expense_ws.append([
                log["date"],
                expense["category"],
                float(expense["amount"]),
                expense["notes"]
            ])

    # --------------------------------------------------------
    # الملخص
    # --------------------------------------------------------

    summary = wb.create_sheet(
        "الملخص"
    )

    summary.sheet_view.rightToLeft = True

    summary.append([
        "التاريخ",
        "عدد العمال",
        "الساعات",
        "الأجور",
        "الفطور",
        "المصاريف الأخرى",
        "الإجمالي"
    ])

    for cell in summary[1]:

        cell.font = Font(
            bold=True
        )

    for log in logs:

        wages = sum(
            float(w["hours"])
            * FIXED_HOURLY_RATE
            for w in log["workers"]
        )

        breakfast = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            in ["فطور", "الفطور"]
        )

        other = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            not in ["فطور", "الفطور"]
        )

        summary.append([
            log["date"],
            len(log["workers"]),
            sum(
                float(w["hours"])
                for w in log["workers"]
            ),
            wages,
            breakfast,
            other,
            wages + breakfast + other
        ])

    # --------------------------------------------------------
    # تنسيق
    # --------------------------------------------------------

    for sheet in wb.worksheets:

        sheet.freeze_panes = "A2"

        for column in sheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                try:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

                except Exception:
                    pass

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40
            )

    filename = "report.xlsx"

    wb.save(filename)

    with open(
        filename,
        "rb"
    ) as file:

        content = file.read()

    os.remove(filename)

    return content


# ============================================================
# PDF
# ============================================================

def generate_pdf_report(logs):

    rows_html = ""

    grand_wages = 0
    grand_breakfast = 0
    grand_other = 0

    for log in logs:

        wages = sum(
            float(w["hours"])
            * FIXED_HOURLY_RATE
            for w in log["workers"]
        )

        breakfast = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            in ["فطور", "الفطور"]
        )

        other = sum(
            float(e["amount"])
            for e in log["expenses"]
            if (
                e["category"]
                or ""
            ).strip().lower()
            not in ["فطور", "الفطور"]
        )

        grand_wages += wages
        grand_breakfast += breakfast
        grand_other += other

        for worker in log["workers"]:

            hours = float(
                worker["hours"]
            )

            wage = (
                hours
                * FIXED_HOURLY_RATE
            )

            rows_html += f"""
            <tr>
                <td>{html.escape(log["date"])}</td>
                <td>{html.escape(worker["name"])}</td>
                <td>{hours:g}</td>
                <td>{FIXED_HOURLY_RATE:.2f}</td>
                <td>{wage:.2f}</td>
            </tr>
            """

    total = (
        grand_wages
        + grand_breakfast
        + grand_other
    )

    document = f"""
    <!DOCTYPE html>

    <html lang="ar" dir="rtl">

    <head>

        <meta charset="UTF-8">

        <style>

        @page {{
            size: A4 landscape;
            margin: 12mm;
        }}

        body {{
            font-family:
                "DejaVu Sans",
                Arial,
                sans-serif;

            direction: rtl;
            padding: 20px;
        }}

        h1 {{
            text-align: center;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 25px;
        }}

        th,
        td {{
            border: 1px solid #999;
            padding: 8px;
            text-align: center;
        }}

        th {{
            background: #eeeeee;
        }}

        .summary {{
            margin-top: 25px;
            font-size: 17px;
            line-height: 2;
        }}

        </style>

    </head>

    <body>

        <h1>
            تقرير إدارة المزرعة
        </h1>

        <p>
            أجر الساعة الثابت:
            <strong>1.50 د.أ</strong>
        </p>

        <table>

            <tr>
                <th>التاريخ</th>
                <th>العامل</th>
                <th>الساعات</th>
                <th>أجر الساعة</th>
                <th>الأجر</th>
            </tr>

            {rows_html}

        </table>

        <div class="summary">

            <strong>
                إجمالي أجور العمال:
                {grand_wages:.2f} د.أ
            </strong>

            <br>

            <strong>
                إجمالي الفطور:
                {grand_breakfast:.2f} د.أ
            </strong>

            <br>

            <strong>
                إجمالي المصاريف الأخرى:
                {grand_other:.2f} د.أ
            </strong>

            <br>

            <strong>
                الإجمالي النهائي:
                {total:.2f} د.أ
            </strong>

        </div>

    </body>

    </html>
    """

    return HTML(
        string=document
    ).write_pdf()


# ============================================================
# BUILD CONFIRMATION DATA
# ============================================================

def build_confirmation_data(
    ai_data
):

    names = clean_names(
        ai_data.get(
            "workers_names",
            []
        )
    )

    count = ai_data.get(
        "workers_count"
    )

    hours = ai_data.get(
        "total_hours"
    )

    breakfast_per_worker = ai_data.get(
        "breakfast_per_worker"
    )

    breakfast_total = ai_data.get(
        "breakfast_total"
    )

    other_expenses = ai_data.get(
        "other_expenses"
    )

    # --------------------------------------------------------
    # إذا الأسماء موجودة والعدد غير موجود
    # --------------------------------------------------------

    if names and count is None:

        count = len(names)

    # --------------------------------------------------------
    # حساب الفطور
    # --------------------------------------------------------

    breakfast_amount = 0.0

    if breakfast_per_worker is not None:

        if count is not None:

            breakfast_amount = (
                float(count)
                * float(
                    breakfast_per_worker
                )
            )

    elif breakfast_total is not None:

        breakfast_amount = float(
            breakfast_total
        )

    # --------------------------------------------------------
    # المصاريف
    # --------------------------------------------------------

    if other_expenses is None:
        other_expenses = 0.0

    # --------------------------------------------------------
    # العمال
    # --------------------------------------------------------

    workers = []

    if names:

        for name in names:

            # إنشاء العامل إذا لم يكن موجوداً
            existing = get_worker_by_name(
                name
            )

            if not existing:
                create_worker(name)

            workers.append({
                "name": name,
                "hours": hours,
                "hourly_rate":
                    FIXED_HOURLY_RATE,
                "task": "",
                "notes": ""
            })

    expenses = []

    if breakfast_amount > 0:

        expenses.append({
            "category": "فطور",
            "amount": breakfast_amount,
            "notes": (
                f"بدل فطور "
                f"{breakfast_per_worker:.2f} د.أ لكل عامل"
                if breakfast_per_worker
                is not None
                else "فطور إجمالي"
            )
        })

    if other_expenses > 0:

        expenses.append({
            "category": "عام",
            "amount": float(
                other_expenses
            ),
            "notes": ""
        })

    return {
        "date":
            ai_data.get("date")
            or today_str(),

        "workers_count":
            count,

        "workers":
            workers,

        "pending_hours":
            hours,

        "breakfast_per_worker":
            breakfast_per_worker,

        "breakfast_total":
            breakfast_total,

        "expenses":
            expenses,

        "notes":
            ai_data.get("notes")
            or ""
    }


# ============================================================
# CONFIRMATION
# ============================================================

def build_confirmation_message(
    data
):

    lines = []

    date_value = data.get(
        "date"
    ) or today_str()

    lines.append(
        f"📋 مراجعة اليومية"
    )

    lines.append(
        f"📅 التاريخ: {date_value}"
    )

    lines.append("")

    workers = data.get(
        "workers",
        []
    )

    count = data.get(
        "workers_count"
    )

    if workers:

        lines.append(
            f"👷 عدد العمال: "
            f"{len(workers)}"
        )

        lines.append("")

        for i, worker in enumerate(
            workers,
            1
        ):

            hours = (
                worker.get("hours")
                or 0
            )

            wage = (
                float(hours)
                * FIXED_HOURLY_RATE
            )

            lines.append(
                f"{i}. {worker['name']} — "
                f"{hours:g} ساعة × "
                f"{FIXED_HOURLY_RATE:.2f} = "
                f"{wage:.2f} د.أ"
            )

    else:

        lines.append(
            f"👷 عدد العمال: "
            f"{count or 'غير محدد'}"
        )

    wages = sum(
        float(
            w.get("hours") or 0
        )
        * FIXED_HOURLY_RATE
        for w in workers
    )

    breakfast = sum(
        float(e["amount"])
        for e in data.get(
            "expenses",
            []
        )
        if (
            e.get("category")
            or ""
        ).strip().lower()
        in ["فطور", "الفطور"]
    )

    other = sum(
        float(e["amount"])
        for e in data.get(
            "expenses",
            []
        )
        if (
            e.get("category")
            or ""
        ).strip().lower()
        not in ["فطور", "الفطور"]
    )

    total = (
        wages
        + breakfast
        + other
    )

    lines.append("")

    lines.append(
        f"⏱️ أجر الساعة الثابت: "
        f"{FIXED_HOURLY_RATE:.2f} د.أ"
    )

    lines.append(
        f"💵 أجور العمال: "
        f"{wages:.2f} د.أ"
    )

    lines.append(
        f"🍳 الفطور: "
        f"{breakfast:.2f} د.أ"
    )

    lines.append(
        f"💰 المصاريف الأخرى: "
        f"{other:.2f} د.أ"
    )

    lines.append(
        f"🏆 الإجمالي: "
        f"{total:.2f} د.أ"
    )

    if data.get("notes"):

        lines.append("")

        lines.append(
            f"📝 ملاحظات: "
            f"{data['notes']}"
        )

    lines.append("")

    # إذا كان هناك سجل سابق
    if daily_exists(date_value):

        lines.append(
            "⚠️ توجد يومية محفوظة مسبقاً "
            f"بتاريخ {date_value}."
        )

        lines.append(
            "لن أستبدلها تلقائياً."
        )

        lines.append("")

    lines.append(
        "هل البيانات صحيحة؟"
    )

    lines.append("")

    lines.append(
        "✅ نعم / حفظ"
    )

    lines.append(
        "✏️ تعديل"
    )

    lines.append(
        "❌ إلغاء"
    )

    return "\n".join(lines)


# ============================================================
# EDIT
# ============================================================

async def edit_pending_data(
    chat_id,
    text,
    data
):

    t = normalize_text(
        text
    )

    # --------------------------------------------------------
    # التاريخ
    # --------------------------------------------------------

    parsed_date = parse_date_text(
        text
    )

    if parsed_date and (
        "تاريخ" in t
        or "اليوم" in t
        or "امس" in t
        or "امبارح" in t
        or "غدا" in t
        or "بكره" in t
        or "بكرة" in t
        or re.search(
            r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}",
            t
        )
    ):

        data["date"] = parsed_date

        save_session(
            chat_id,
            "AWAITING_CONFIRMATION",
            data
        )

        await send_message(
            chat_id,
            build_confirmation_message(
                data
            )
        )

        return True

    # --------------------------------------------------------
    # الساعات
    # --------------------------------------------------------

    match = re.search(
        r"(?:الساعات|ساعات|ساعة|الساعه)"
        r"\s*(?:=|:)?\s*"
        r"(\d+(?:[.,]\d+)?)",
        t,
        flags=re.IGNORECASE
    )

    if match:

        hours = float(
            match.group(1).replace(
                ",",
                "."
            )
        )

        if not (
            0 <= hours <= 24
        ):

            await send_message(
                chat_id,
                "⚠️ الساعات يجب أن تكون بين 0 و24."
            )

            return True

        for worker in data.get(
            "workers",
            []
        ):

            worker["hours"] = hours

        data["pending_hours"] = hours

        save_session(
            chat_id,
            "AWAITING_CONFIRMATION",
            data
        )

        await send_message(
            chat_id,
            build_confirmation_message(
                data
            )
        )

        return True

    # --------------------------------------------------------
    # فطور لكل عامل
    # --------------------------------------------------------

    match = re.search(
        r"فطور\s+لكل\s+عامل\s*"
        r"(?:=|:)?\s*"
        r"(\d+(?:[.,]\d+)?)",
        t,
        flags=re.IGNORECASE
    )

    if match:

        per_worker = float(
            match.group(1).replace(
                ",",
                "."
            )
        )

        count = data.get(
            "workers_count"
        )

        if not count:
            count = len(
                data.get(
                    "workers",
                    []
                )
            )

        amount = (
            count
            * per_worker
        )

        data[
            "breakfast_per_worker"
        ] = per_worker

        data["expenses"] = [
            e
            for e in data.get(
                "expenses",
                []
            )
            if e.get("category")
            not in ["فطور", "الفطور"]
        ]

        data["expenses"].insert(
            0,
            {
                "category": "فطور",
                "amount": amount,
                "notes":
                    f"{per_worker:.2f} د.أ لكل عامل"
            }
        )

        save_session(
            chat_id,
            "AWAITING_CONFIRMATION",
            data
        )

        await send_message(
            chat_id,
            build_confirmation_message(
                data
            )
        )

        return True

    # --------------------------------------------------------
    # المصاريف
    # --------------------------------------------------------

    match = re.search(
        r"(?:المصاريف|مصاريف|صرف|دفعت)"
        r"\s*(?:=|:)?\s*"
        r"(\d+(?:[.,]\d+)?)",
        t,
        flags=re.IGNORECASE
    )

    if match:

        amount = float(
            match.group(1).replace(
                ",",
                "."
            )
        )

        if amount < 0:

            await send_message(
                chat_id,
                "⚠️ المصروف لا يمكن أن يكون سالباً."
            )

            return True

        data["expenses"] = [
            e
            for e in data.get(
                "expenses",
                []
            )
            if e.get("category")
            in ["فطور", "الفطور"]
        ]

        if amount > 0:

            data["expenses"].append({
                "category": "عام",
                "amount": amount,
                "notes": ""
            })

        save_session(
            chat_id,
            "AWAITING_CONFIRMATION",
            data
        )

        await send_message(
            chat_id,
            build_confirmation_message(
                data
            )
        )

        return True

    # --------------------------------------------------------
    # إضافة عامل
    # --------------------------------------------------------

    match = re.search(
        r"(?:أضف|اضف)\s+(?:العامل\s+)?(.+)",
        text,
        flags=re.IGNORECASE
    )

    if match:

        name = normalize_name(
            match.group(1)
        )

        if validate_worker_name(name):

            existing = get_worker_by_name(
                name
            )

            if not existing:
                create_worker(name)

            data.setdefault(
                "workers",
                []
            ).append({
                "name": name,
                "hours":
                    data.get(
                        "pending_hours"
                    ),
                "hourly_rate":
                    FIXED_HOURLY_RATE,
                "task": "",
                "notes": ""
            })

            data["workers_count"] = len(
                data["workers"]
            )

            # إعادة حساب الفطور لكل العمال
            per_worker = data.get(
                "breakfast_per_worker"
            )

            if per_worker is not None:

                for expense in data.get(
                    "expenses",
                    []
                ):

                    if expense.get(
                        "category"
                    ) == "فطور":

                        expense["amount"] = (
                            len(data["workers"])
                            * float(per_worker)
                        )

            save_session(
                chat_id,
                "AWAITING_CONFIRMATION",
                data
            )

            await send_message(
                chat_id,
                build_confirmation_message(
                    data
                )
            )

            return True

    return False


# ============================================================
# SESSION HANDLER
# ============================================================

async def handle_active_session(
    chat_id,
    text,
    session
):

    step = session["step"]
    data = session["data"]

    # --------------------------------------------------------
    # CONFIRMATION
    # --------------------------------------------------------

    if step == "AWAITING_CONFIRMATION":

        if is_yes(text):

            if not data.get(
                "workers"
            ):

                await send_message(
                    chat_id,
                    "⚠️ لا يمكن الحفظ قبل تحديد أسماء العمال."
                )

                return True

            missing_hours = [
                w["name"]
                for w in data["workers"]
                if w.get("hours") is None
            ]

            if missing_hours:

                await send_message(
                    chat_id,
                    "⚠️ لم يتم تحديد ساعات العمل لـ:\n"
                    + "\n".join(
                        missing_hours
                    )
                )

                return True

            if daily_exists(
                data["date"]
            ):

                await send_message(
                    chat_id,
                    f"⚠️ توجد يومية محفوظة بالفعل "
                    f"بتاريخ {data['date']}.\n\n"
                    "لن أستبدلها تلقائياً.\n"
                    "إذا كنت تريد تسجيل يوم جديد، "
                    "غيّر التاريخ ثم أعد التأكيد."
                )

                clear_session(
                    chat_id
                )

                return True

            try:

                save_daily_data(
                    data
                )

            except Exception as e:

                logger.exception(
                    "Save failed: %s",
                    e
                )

                await send_message(
                    chat_id,
                    "❌ حدث خطأ أثناء الحفظ. "
                    "لم يتم حفظ أي جزء من البيانات."
                )

                return True

            clear_session(
                chat_id
            )

            log = get_daily_log(
                data["date"]
            )

            await send_message(
                chat_id,
                "✅ تم حفظ اليومية بنجاح.\n\n"
                + format_daily_summary(
                    log
                )
            )

            return True

        if normalize_text(
            text
        ).lower().strip() in {
            "الغاء",
            "إلغاء",
            "الغاء العملية",
            "لا تحفظ"
        }:

            clear_session(
                chat_id
            )

            await send_message(
                chat_id,
                "❌ تم إلغاء العملية."
            )

            return True

        if normalize_text(
            text
        ).lower().strip() in {
            "تعديل",
            "عدل"
        }:

            save_session(
                chat_id,
                "EDITING",
                data
            )

            await send_message(
                chat_id,
                """
✏️ ماذا تريد تعديل؟

مثال:

التاريخ أمس

الساعات 7

فطور لكل عامل 1

المصاريف 15

أضف العامل أحمد

اكتب التعديل مباشرة.
"""
            )

            return True

        await send_message(
            chat_id,
            "اكتب: نعم للحفظ، تعديل للتعديل، أو إلغاء."
        )

        return True

    # --------------------------------------------------------
    # EDITING
    # --------------------------------------------------------

    if step == "EDITING":

        if await edit_pending_data(
            chat_id,
            text,
            data
        ):

            return True

        await send_message(
            chat_id,
            "لم أفهم التعديل.\n\n"
            "مثال: الساعات 7\n"
            "أو: فطور لكل عامل 1\n"
            "أو: المصاريف 15\n"
            "أو: التاريخ أمس"
        )

        return True

    # --------------------------------------------------------
    # COUNT
    # --------------------------------------------------------

    if step == "ASK_WORKERS_COUNT":

        count = number_from_text(
            text
        )

        if count is None:

            await send_message(
                chat_id,
                "اكتب عدد العمال كرقم، مثل: 5"
            )

            return True

        count = int(count)

        if not (
            1 <= count <= MAX_WORKERS
        ):

            await send_message(
                chat_id,
                f"عدد العمال يجب أن يكون بين 1 و{MAX_WORKERS}."
            )

            return True

        data[
            "workers_count"
        ] = count

        save_session(
            chat_id,
            "ASK_NAMES",
            data
        )

        await send_message(
            chat_id,
            f"👷 تم تسجيل {count} عمال.\n"
            "اكتب أسماءهم مفصولة بفواصل."
        )

        return True

    # --------------------------------------------------------
    # NAMES
    # --------------------------------------------------------

    if step == "ASK_NAMES":

        names = clean_names(
            re.split(
                r"[,،;\n]+|\s+و\s+",
                text
            )
        )

        expected = data.get(
            "workers_count"
        )

        if len(names) != expected:

            await send_message(
                chat_id,
                f"⚠️ قلت إن عدد العمال {expected}، "
                f"لكن وجدت {len(names)} أسماء.\n\n"
                "لن أخمّن.\n"
                "اكتب الأسماء مفصولة بفواصل."
            )

            return True

        data["workers"] = []

        for name in names:

            if not get_worker_by_name(
                name
            ):
                create_worker(name)

            data["workers"].append({
                "name": name,
                "hours": None,
                "hourly_rate":
                    FIXED_HOURLY_RATE,
                "task": "",
                "notes": ""
            })

        save_session(
            chat_id,
            "ASK_HOURS",
            data
        )

        await send_message(
            chat_id,
            "⏱️ كم ساعة عمل لكل عامل؟\n\n"
            "إذا كانت الساعات متساوية اكتب رقماً واحداً.\n"
            "مثال: 8\n\n"
            "وإذا كانت مختلفة:\n"
            "أحمد 8، محمد 7، خالد 8"
        )

        return True

    # --------------------------------------------------------
    # HOURS
    # --------------------------------------------------------

    if step == "ASK_HOURS":

        # رقم واحد
        if re.fullmatch(
            r"\s*\d+(?:[.,]\d+)?\s*",
            normalize_text(text)
        ):

            hours = number_from_text(
                text
            )

            if not (
                0 <= hours <= 24
            ):

                await send_message(
                    chat_id,
                    "⚠️ الساعات يجب أن تكون بين 0 و24."
                )

                return True

            for worker in data["workers"]:
                worker["hours"] = hours

            data[
                "pending_hours"
            ] = hours

        else:

            # اسم + رقم
            pairs = re.findall(
                r"([^,\n]+?)\s+(\d+(?:[.,]\d+)?)",
                normalize_text(text)
            )

            assigned = {}

            for raw_name, raw_hours in pairs:

                name = normalize_name(
                    raw_name
                )

                hours = float(
                    raw_hours.replace(
                        ",",
                        "."
                    )
                )

                if not (
                    0 <= hours <= 24
                ):

                    await send_message(
                        chat_id,
                        "⚠️ الساعات يجب أن تكون بين 0 و24."
                    )

                    return True

                assigned[
                    name.lower()
                ] = hours

            missing = []

            for worker in data["workers"]:

                key = worker[
                    "name"
                ].lower()

                if key in assigned:

                    worker["hours"] = (
                        assigned[key]
                    )

                else:

                    missing.append(
                        worker["name"]
                    )

            if missing:

                await send_message(
                    chat_id,
                    "لم أحدد ساعات:\n"
                    + "\n".join(missing)
                )

                return True

        save_session(
            chat_id,
            "ASK_EXPENSES",
            data
        )

        await send_message(
            chat_id,
            """
🍳 هل يوجد بدل فطور؟

إذا كان لكل عامل:
"فطور لكل عامل 1"

إذا كان مبلغاً إجمالياً:
"فطور 6 دنانير"

وإذا لا يوجد فطور:
"لا"
"""
        )

        return True

    # --------------------------------------------------------
    # BREAKFAST
    # --------------------------------------------------------

    if step == "ASK_EXPENSES":

        t = normalize_text(
            text
        )

        # بدون فطور
        if is_no(text):

            data[
                "breakfast_per_worker"
            ] = None

            data[
                "breakfast_total"
            ] = 0

            data["expenses"] = []

            save_session(
                chat_id,
                "ASK_OTHER_EXPENSES",
                data
            )

            await send_message(
                chat_id,
                "💰 هل توجد مصاريف أخرى؟\n"
                "اكتب المبلغ أو «لا»."
            )

            return True

        # فطور لكل عامل
        match = re.search(
            r"فطور\s+لكل\s+عامل\s*"
            r"(?:=|:)?\s*"
            r"(\d+(?:[.,]\d+)?)",
            t,
            flags=re.IGNORECASE
        )

        if match:

            per_worker = float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

            if not validate_amount(
                per_worker,
                MAX_BREAKFAST_PER_WORKER
            ):

                await send_message(
                    chat_id,
                    "⚠️ قيمة الفطور غير صحيحة."
                )

                return True

            count = data[
                "workers_count"
            ]

            breakfast_total = (
                count
                * per_worker
            )

            data[
                "breakfast_per_worker"
            ] = per_worker

            data[
                "breakfast_total"
            ] = breakfast_total

            data["expenses"] = [{
                "category": "فطور",
                "amount": breakfast_total,
                "notes":
                    f"{per_worker:.2f} د.أ لكل عامل"
            }]

            save_session(
                chat_id,
                "ASK_OTHER_EXPENSES",
                data
            )

            await send_message(
                chat_id,
                f"🍳 الفطور: "
                f"{count} × "
                f"{per_worker:.2f} = "
                f"{breakfast_total:.2f} د.أ\n\n"
                "💰 هل توجد مصاريف أخرى؟\n"
                "اكتب المبلغ أو «لا»."
            )

            return True

        # فطور إجمالي
        match = re.search(
            r"(?:فطور|الفطور)\s*"
            r"(?:=|:)?\s*"
            r"(\d+(?:[.,]\d+)?)",
            t,
            flags=re.IGNORECASE
        )

        if match:

            total = float(
                match.group(1).replace(
                    ",",
                    "."
                )
            )

            if not validate_amount(
                total
            ):

                await send_message(
                    chat_id,
                    "⚠️ قيمة الفطور غير صحيحة."
                )

                return True

            data[
                "breakfast_total"
            ] = total

            data[
                "breakfast_per_worker"
            ] = None

            data["expenses"] = [{
                "category": "فطور",
                "amount": total,
                "notes":
                    "فطور إجمالي"
            }]

            save_session(
                chat_id,
                "ASK_OTHER_EXPENSES",
                data
            )

            await send_message(
                chat_id,
                f"🍳 تم تسجيل فطور إجمالي "
                f"{total:.2f} د.أ.\n\n"
                "💰 هل توجد مصاريف أخرى؟\n"
                "اكتب المبلغ أو «لا»."
            )

            return True

        await send_message(
            chat_id,
            "لم أفهم الفطور.\n\n"
            "مثال:\n"
            "فطور لكل عامل 1\n\n"
            "أو:\n"
            "فطور 6 دنانير\n\n"
            "أو:\n"
            "لا"
        )

        return True

    # --------------------------------------------------------
    # OTHER EXPENSES
    # --------------------------------------------------------

    if step == "ASK_OTHER_EXPENSES":

        if is_no(text):

            data["expenses"] = [
                e
                for e in data.get(
                    "expenses",
                    []
                )
                if e.get("category")
                in ["فطور", "الفطور"]
            ]

        else:

            amount = number_from_text(
                text
            )

            if amount is None:

                await send_message(
                    chat_id,
                    "اكتب مبلغ المصاريف أو «لا»."
                )

                return True

            if not (
                0 <= amount <= MAX_EXPENSE
            ):

                await send_message(
                    chat_id,
                    "⚠️ قيمة المصاريف غير صحيحة."
                )

                return True

            data["expenses"] = [
                e
                for e in data.get(
                    "expenses",
                    []
                )
                if e.get("category")
                in ["فطور", "الفطور"]
            ]

            if amount > 0:

                data["expenses"].append({
                    "category": "عام",
                    "amount": amount,
                    "notes": ""
                })

        save_session(
            chat_id,
            "ASK_NOTES",
            data
        )

        await send_message(
            chat_id,
            "📝 هل توجد ملاحظات؟\n"
            "اكتبها أو «لا»."
        )

        return True

    # --------------------------------------------------------
    # NOTES
    # --------------------------------------------------------

    if step == "ASK_NOTES":

        if is_no(text):

            data["notes"] = ""

        else:

            data["notes"] = (
                text.strip()[:1000]
            )

        save_session(
            chat_id,
            "AWAITING_CONFIRMATION",
            data
        )

        await send_message(
            chat_id,
            build_confirmation_message(
                data
            )
        )

        return True

    return False


# ============================================================
# COMMANDS
# ============================================================

async def handle_command(
    chat_id,
    text
):

    normalized = normalize_text(
        text
    ).lower().strip()

    # --------------------------------------------------------
    # CANCEL
    # --------------------------------------------------------

    if normalized in {
        "/cancel",
        "الغاء",
        "إلغاء",
        "الغاء العملية"
    }:

        clear_session(
            chat_id
        )

        await send_message(
            chat_id,
            "❌ تم إلغاء العملية."
        )

        return True

    # --------------------------------------------------------
    # HELP
    # --------------------------------------------------------

    if normalized in {
        "/start",
        "/help",
        "مساعدة"
    }:

        clear_session(
            chat_id
        )

        await send_message(
            chat_id,
            """
🤖 أهلاً بك في Medjol Farm Manager V3

أرسل بيانات العمل بطريقتك الطبيعية.

مثال كامل:

"اليوم اشتغل أحمد ومحمد وخالد 8 ساعات، فطور لكل عامل دينار، ومصاريف 10 دنانير"

سأفهم تلقائياً:

📅 التاريخ
👷 العمال
⏱️ ساعات العمل
💵 أجر الساعة = 1.50 د.أ ثابت
🍳 الفطور
💰 المصاريف
📝 الملاحظات

مثال آخر:

"أمس كان عندي 5 عمال، كل واحد 7 ساعات، الفطور لكل عامل 1 دينار، والبنزين 5 دنانير"

الأوامر:

📋 تقرير
📅 تقرير اليوم
📆 تقرير الأسبوع
🗓️ تقرير الشهر
📄 تقرير PDF
📗 تقرير Excel

👷 العمال
➕ أضف العامل أحمد
❌ احذف العامل أحمد

❌ إلغاء
"""
        )

        return True

    # --------------------------------------------------------
    # WORKERS
    # --------------------------------------------------------

    if normalized in {
        "العمال",
        "قائمة العمال",
        "/workers"
    }:

        workers = get_active_workers()

        if not workers:

            await send_message(
                chat_id,
                "لا يوجد عمال مسجلون."
            )

            return True

        lines = [
            "👷 العمال المسجلون:",
            ""
        ]

        for i, worker in enumerate(
            workers,
            1
        ):

            lines.append(
                f"{i}. {worker['name']} "
                f"— {FIXED_HOURLY_RATE:.2f} د.أ/ساعة"
            )

        await send_message(
            chat_id,
            "\n".join(lines)
        )

        return True

    # --------------------------------------------------------
    # ADD WORKER
    # --------------------------------------------------------

    if (
        normalized.startswith(
            "اضف العامل "
        )
        or normalized.startswith(
            "أضف العامل "
        )
        or normalized.startswith(
            "/addworker "
        )
    ):

        if normalized.startswith(
            "/addworker "
        ):

            name = text.split(
                " ",
                1
            )[1]

        else:

            name = re.sub(
                r"^(اضف|أضف)\s+العامل\s+",
                "",
                text,
                flags=re.IGNORECASE
            )

        name = normalize_name(
            name
        )

        if not validate_worker_name(
            name
        ):

            await send_message(
                chat_id,
                "⚠️ اسم العامل غير واضح."
            )

            return True

        if get_worker_by_name(
            name
        ):

            await send_message(
                chat_id,
                f"⚠️ العامل {name} موجود بالفعل."
            )

            return True

        create_worker(
            name
        )

        await send_message(
            chat_id,
            f"✅ تم إضافة العامل: {name}"
        )

        return True

    # --------------------------------------------------------
    # DELETE WORKER
    # --------------------------------------------------------

    if (
        normalized.startswith(
            "احذف العامل "
        )
        or normalized.startswith(
            "حذف العامل "
        )
    ):

        name = re.sub(
            r"^(احذف|حذف)\s+العامل\s+",
            "",
            text,
            flags=re.IGNORECASE
        )

        name = normalize_name(
            name
        )

        if deactivate_worker(
            name
        ):

            await send_message(
                chat_id,
                f"✅ تم تعطيل العامل: {name}"
            )

        else:

            await send_message(
                chat_id,
                f"⚠️ لم أجد العامل: {name}"
            )

        return True

    # --------------------------------------------------------
    # RATE - ممنوع تغييره
    # --------------------------------------------------------

    if re.search(
        r"(?:أجر|اجر|سعر)\s*(?:الساعة|الساعه)",
        normalized
    ):

        await send_message(
            chat_id,
            "ℹ️ أجر الساعة ثابت في النظام: "
            "1.50 د.أ لكل ساعة."
        )

        return True

    return False


# ============================================================
# REPORT COMMANDS
# ============================================================

async def handle_report_command(
    chat_id,
    text
):

    t = normalize_text(
        text
    ).lower().strip()

    is_report = (
        t.startswith("تقرير")
        or t in {
            "/report",
            "report"
        }
    )

    if not is_report:
        return False

    today = now_local().date()

    # --------------------------------------------------------
    # التاريخ
    # --------------------------------------------------------

    if "اليوم" in t:

        date_from = today
        date_to = today

    elif (
        "اسبوع" in t
        or "أسبوع" in t
    ):

        date_from = (
            today
            - timedelta(
                days=today.weekday()
            )
        )

        date_to = today

    elif (
        "شهر" in t
        or "شهري" in t
    ):

        date_from = today.replace(
            day=1
        )

        date_to = today

    else:

        date_from = date(
            2000,
            1,
            1
        )

        date_to = today

    logs = get_logs_between(
        date_from.isoformat(),
        date_to.isoformat()
    )

    # --------------------------------------------------------
    # PDF
    # --------------------------------------------------------

    if "pdf" in t:

        if not logs:

            await send_message(
                chat_id,
                "لا توجد بيانات لإنشاء التقرير."
            )

            return True

        try:

            pdf = generate_pdf_report(
                logs
            )

            await send_document(
                chat_id,
                "تقرير_المزرعة.pdf",
                pdf,
                "📄 تقرير المزرعة"
            )

        except Exception as e:

            logger.exception(
                "PDF failed: %s",
                e
            )

            await send_message(
                chat_id,
                "❌ تعذر إنشاء ملف PDF."
            )

        return True

    # --------------------------------------------------------
    # EXCEL
    # --------------------------------------------------------

    if (
        "excel" in t
        or "اكسل" in t
        or "إكسل" in t
    ):

        if not logs:

            await send_message(
                chat_id,
                "لا توجد بيانات لإنشاء التقرير."
            )

            return True

        try:

            excel = generate_excel_report(
                logs
            )

            await send_document(
                chat_id,
                "تقرير_المزرعة.xlsx",
                excel,
                "📗 تقرير المزرعة"
            )

        except Exception as e:

            logger.exception(
                "Excel failed: %s",
                e
            )

            await send_message(
                chat_id,
                "❌ تعذر إنشاء ملف Excel."
            )

        return True

    # --------------------------------------------------------
    # TEXT
    # --------------------------------------------------------

    if date_from == date_to:

        title = (
            f"📊 تقرير يوم "
            f"{date_from.isoformat()}"
        )

    else:

        title = (
            f"📊 التقرير من "
            f"{date_from.isoformat()} "
            f"إلى "
            f"{date_to.isoformat()}"
        )

    report = generate_text_report(
        logs,
        title
    )

    await send_message(
        chat_id,
        report
    )

    return True


# ============================================================
# MAIN MESSAGE HANDLER
# ============================================================

async def handle_incoming_message(
    chat_id,
    text
):

    text = text.strip()

    if not text:
        return

    # --------------------------------------------------------
    # COMMANDS
    # --------------------------------------------------------

    if await handle_command(
        chat_id,
        text
    ):
        return

    # --------------------------------------------------------
    # SESSION
    # --------------------------------------------------------

    session = get_session(
        chat_id
    )

    if session:

        await handle_active_session(
            chat_id,
            text,
            session
        )

        return

    # --------------------------------------------------------
    # REPORTS
    # --------------------------------------------------------

    if await handle_report_command(
        chat_id,
        text
    ):
        return

    # --------------------------------------------------------
    # AI
    # --------------------------------------------------------

    await send_message(
        chat_id,
        "⏳ جارٍ فهم البيانات..."
    )

    ai_data = await extract_data_with_ai(
        text
    )

    # --------------------------------------------------------
    # FALLBACK
    # --------------------------------------------------------

    if not ai_data:

        ai_data = deterministic_extract(
            text
        )

    if not ai_data:

        await send_message(
            chat_id,
            """
لم أستطع فهم الرسالة بشكل موثوق.

جرب مثلاً:

"اليوم اشتغل أحمد ومحمد 8 ساعات، فطور لكل عامل 1 دينار، ومصاريف 10"

أو:

"أمس 5 عمال، 7 ساعات، الفطور لكل عامل دينار، والبنزين 5 دنانير"
"""
        )

        return

    # --------------------------------------------------------
    # VALIDATION ERROR
    # --------------------------------------------------------

    if ai_data.get(
        "_validation_error"
    ):

        await send_message(
            chat_id,
            "⚠️ "
            + ai_data[
                "_validation_error"
            ]
        )

        return

    # --------------------------------------------------------
    # BUILD DATA
    # --------------------------------------------------------

    data = build_confirmation_data(
        ai_data
    )

    names = data.get(
        "workers",
        []
    )

    count = data.get(
        "workers_count"
    )

    # --------------------------------------------------------
    # NO WORKERS
    # --------------------------------------------------------

    if not count and not names:

        save_session(
            chat_id,
            "ASK_WORKERS_COUNT",
            data
        )

        await send_message(
            chat_id,
            "👷 كم عدد العمال؟"
        )

        return

    # --------------------------------------------------------
    # COUNT BUT NO NAMES
    # --------------------------------------------------------

    if count and not names:

        save_session(
            chat_id,
            "ASK_NAMES",
            data
        )

        await send_message(
            chat_id,
            f"👷 فهمت أن عدد العمال {count}.\n"
            "اكتب أسماء العمال مفصولة بفواصل."
        )

        return

    # --------------------------------------------------------
    # NAMES BUT COUNT CONFLICT
    # --------------------------------------------------------

    if (
        count
        and names
        and len(names) != count
    ):

        await send_message(
            chat_id,
            f"⚠️ ذكرت {count} عمال "
            f"لكن وجدت {len(names)} أسماء.\n"
            "لن أخمّن.\n\n"
            "اكتب العدد الصحيح."
        )

        data[
            "workers_count"
        ] = None

        save_session(
            chat_id,
            "ASK_WORKERS_COUNT",
            data
        )

        return

    # --------------------------------------------------------
    # HOURS MISSING
    # --------------------------------------------------------

    missing_hours = [
        w["name"]
        for w in names
        if w.get("hours") is None
    ]

    if missing_hours:

        save_session(
            chat_id,
            "ASK_HOURS",
            data
        )

        await send_message(
            chat_id,
            "⏱️ كم ساعة عمل لكل عامل؟\n"
            "إذا كانت الساعات متساوية اكتب رقماً واحداً."
        )

        return

    # --------------------------------------------------------
    # إذا لم يذكر فطور أو مصاريف
    # لا نسأله عنها إذا كانت الرسالة مكتملة.
    # تعتبر 0 تلقائياً.
    # --------------------------------------------------------

    save_session(
        chat_id,
        "AWAITING_CONFIRMATION",
        data
    )

    await send_message(
        chat_id,
        build_confirmation_message(
            data
        )
    )


# ============================================================
# TELEGRAM POLLING
# ============================================================

async def delete_webhook():

    if not TELEGRAM_BOT_TOKEN:
        return

    url = (
        f"https://api.telegram.org/"
        f"bot{TELEGRAM_BOT_TOKEN}/deleteWebhook"
    )

    try:

        async with httpx.AsyncClient() as client:

            await client.get(
                url,
                params={
                    "drop_pending_updates": False
                },
                timeout=15
            )

    except Exception as e:

        logger.error(
            "Webhook deletion failed: %s",
            e
        )


async def polling_loop():

    if not TELEGRAM_BOT_TOKEN:

        logger.error(
            "❌ لا يمكن تشغيل Telegram بدون TELEGRAM_BOT_TOKEN"
        )

        return

    await delete_webhook()

    offset = 0

    logger.info(
        "🚀 Medjol Farm Manager V3 started"
    )

    url = (
        f"https://api.telegram.org/"
        f"bot{TELEGRAM_BOT_TOKEN}/getUpdates"
    )

    async with httpx.AsyncClient() as client:

        while True:

            try:

                response = await client.get(
                    url,
                    params={
                        "offset": offset,
                        "timeout": 30,
                        "allowed_updates":
                            json.dumps(
                                ["message"]
                            )
                    },
                    timeout=40
                )

                if response.status_code != 200:

                    logger.error(
                        "Polling HTTP error: %s",
                        response.status_code
                    )

                    await asyncio.sleep(3)

                    continue

                updates = response.json().get(
                    "result",
                    []
                )

                for update in updates:

                    offset = (
                        update["update_id"]
                        + 1
                    )

                    message = update.get(
                        "message"
                    )

                    if not message:
                        continue

                    if "text" not in message:
                        continue

                    chat = message.get(
                        "chat",
                        {}
                    )

                    chat_id = chat.get(
                        "id"
                    )

                    text = message.get(
                        "text",
                        ""
                    ).strip()

                    if not chat_id or not text:
                        continue

                    try:

                        await handle_incoming_message(
                            chat_id,
                            text
                        )

                    except Exception as e:

                        logger.exception(
                            "Message handling error: %s",
                            e
                        )

                        await send_message(
                            chat_id,
                            "❌ حدث خطأ غير متوقع. "
                            "لم يتم حفظ هذه العملية."
                        )

            except asyncio.CancelledError:

                raise

            except Exception as e:

                logger.exception(
                    "Polling exception: %s",
                    e
                )

                await asyncio.sleep(5)


# ============================================================
# FASTAPI
# ============================================================

polling_task = None


@app.on_event("startup")
async def startup():

    global polling_task

    if not TELEGRAM_BOT_TOKEN:

        logger.error(
            "❌ TELEGRAM_BOT_TOKEN غير موجود"
        )

        return

    if (
        polling_task is None
        or polling_task.done()
    ):

        polling_task = asyncio.create_task(
            polling_loop()
        )

    logger.info(
        "✅ Bot polling started"
    )


@app.on_event("shutdown")
async def shutdown():

    global polling_task

    if polling_task:

        polling_task.cancel()

        try:
            await polling_task
        except asyncio.CancelledError:
            pass

        polling_task = None


@app.get("/")
async def root():

    return {
        "status": "running",
        "name": "Medjol Farm Manager",
        "version": "3.0",
        "mode": "polling",
        "hourly_rate": FIXED_HOURLY_RATE,
        "timezone": TIMEZONE
    }


@app.get("/health")
async def health():

    return {
        "status": "healthy",
        "time": now_local().isoformat(),
        "hourly_rate": FIXED_HOURLY_RATE
    }


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    uvicorn.run(
        app,
        host="0.0.0.0",
        port=PORT
    )
